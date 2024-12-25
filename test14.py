import sys
import serial
import json
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QPushButton, QTableWidget, QTableWidgetItem, QMessageBox, QInputDialog,
    QCheckBox, QScrollArea
)
import jdatetime
from PyQt5.QtCore import QTimer
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox
import json
from openpyxl import Workbook, load_workbook

class TeacherManager(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("مدیریت و ثبت حضور اساتید")
        self.setGeometry(200, 100, 1200, 800)
        

        # تعریف روزها و ساعت‌ها
        self.days_of_week = ['دوشنبه', 'سه شنبه', 'چهارشنبه', 'پنجشنبه', 'جمعه', 'شنبه', 'یک شنبه']
        self.time_slots = ['7:30-9:30', '9:30-11:30', '12:00-14:00', '14:00-16:00', '16:00-18:00', '18:00-20:00']


        # متغیر برای ذخیره اطلاعات اساتید
        self.teachers = {}
        
        # زمان آخرین حضور هر استاد
        self.last_attendance = {}

        # چک باکس‌های روزها و ساعت‌ها
        self.days_time_checkboxes = {}
        main_layout = QVBoxLayout()

        # ایجاد یک QScrollArea برای روزها
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)

        # ایجاد یک QWidget برای قرار دادن چک باکس‌ها در آن
        day_widget = QWidget()
        day_layout = QVBoxLayout()



        for day in self.days_of_week:
            day_checkbox = QCheckBox(day)
            self.days_time_checkboxes[day] = (day_checkbox, [])
            day_layout.addWidget(day_checkbox)

            # ایجاد چک باکس‌های ساعت برای هر روز
            time_layout = QVBoxLayout()  # برای قرار دادن ساعت‌ها به صورت عمودی
            for time in self.time_slots:
                time_checkbox = QCheckBox(time)
                time_layout.addWidget(time_checkbox)
                self.days_time_checkboxes[day][1].append(time_checkbox)  # اضافه کردن چک باکس ساعت به لیست

            day_layout.addLayout(time_layout)

        # ایجاد و تنظیم لایه اصلی
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        # تعریف لایه اصلی به درستی
        self.layout = QVBoxLayout()  # تعریف لایه اصلی
        self.central_widget.setLayout(self.layout)  # تنظیم لایه اصلی برای ویجت مرکزی
        # تعریف input_layout
        self.input_layout = QVBoxLayout()  # اضافه کردن لایه ورودی
        self.layout.addLayout(self.input_layout)  # اضافه کردن لایه ورودی به لایه اصلی
        
        day_widget.setLayout(day_layout)
        scroll_area.setWidget(day_widget)  # قرار دادن ویجت روزها در اسکرول

        main_layout.addWidget(scroll_area)  # اضافه کردن اسکرول به layout اصلی

        # ورودی اطلاعات استاد
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("نام استاد")
        self.code_input = QLineEdit()
        self.code_input.setPlaceholderText("کد استاد")
        self.class_input = QLineEdit()
        self.class_input.setPlaceholderText("کلاس استاد")
        self.subject_input = QLineEdit()
        self.subject_input.setPlaceholderText("نام درس")
        self.add_button = QPushButton("افزودن استاد")
        self.add_button.clicked.connect(self.add_teacher)

        input_layout = QHBoxLayout()
        input_layout.addWidget(self.name_input)
        input_layout.addWidget(self.code_input)
        input_layout.addWidget(self.class_input)
        input_layout.addWidget(self.subject_input)
        input_layout.addWidget(self.add_button)
        main_layout.addLayout(input_layout)

        # جدول نمایش اساتید
        self.teacher_table = QTableWidget()
        self.teacher_table.setColumnCount(5)
        self.teacher_table.setHorizontalHeaderLabels(["کد استاد", "نام استاد", "کلاس","نام درس", "روزها و زمان"])
        main_layout.addWidget(self.teacher_table)

        # جدول ثبت حضور
        self.attendance_table = QTableWidget()
        self.attendance_table.setColumnCount(6)
        self.attendance_table.setHorizontalHeaderLabels(["زمان", "کد", "نام", "کلاس", "روز", "زمان کلاس"])
        main_layout.addWidget(self.attendance_table)

        # دکمه شروع و توقف دریافت داده
        self.start_button = QPushButton("شروع ثبت حضور")
        self.start_button.clicked.connect(self.start_receiving_data)
        main_layout.addWidget(self.start_button)
        

        # دکمه ویرایش استاد
        self.edit_button = QPushButton("ویرایش استاد")
        self.edit_button.clicked.connect(self.edit_teacher)
        main_layout.addWidget(self.edit_button)

        # دکمه حذف استاد
        self.delete_button = QPushButton("حذف استاد")
        self.delete_button.clicked.connect(self.delete_teacher)
        main_layout.addWidget(self.delete_button)
        self.save_attendance_button = QPushButton("ذخیره حضور در فایل Excel")
        self.save_attendance_button.clicked.connect(self.save_attendance_to_excel)
        main_layout.addWidget(self.save_attendance_button)

        self.teachers = {}
        self.load_teachers_from_file() 
        # تنظیمات پورت سریال
        self.port = 'COM5'
        self.baudrate = 9600
        try:
            self.ser = serial.Serial(self.port, self.baudrate, timeout=1)
            print(f"Connected to {self.port} at {self.baudrate} baud.")
        except serial.SerialException as e:
            QMessageBox.critical(self, "خطا", f"اتصال به پورت سریال امکان‌پذیر نیست: {e}")
            self.ser = None

        # بارگذاری اطلاعات اساتید از فایل JSON
        self.load_teachers_from_file()

        # اضافه کردن ویجت مرکزی
        central_widget = QWidget()  # ایجاد یک ویجت مرکزی
        central_widget.setLayout(main_layout)  # تنظیم layout برای ویجت مرکزی
        self.setCentralWidget(central_widget)  # تنظیم ویجت مرکزی برای پنجره اصلی
        QApplication.instance().aboutToQuit.connect(self.save_teachers_to_file)  # ذخیره‌سازی اطلاعات اساتید هنگام بسته شدن
        # فرض کنید دکمه‌ای دارید که این تابع را فراخوانی می‌کند
            
    def add_teacher(self):
        name = self.name_input.text().strip()
        code = self.code_input.text().strip()
        class_name = self.class_input.text().strip()
        course = self.subject_input.text().strip()

        if not all([name, code, class_name, course]):
            QMessageBox.warning(self, "خطا", "لطفاً تمامی فیلدها را پر کنید.")
            return

        selected_days = []
        for day, (day_checkbox, time_checkboxes) in self.days_time_checkboxes.items():
            if day_checkbox.isChecked():
                selected_times = [time.text() for time in time_checkboxes if time.isChecked()]
                if selected_times:
                    selected_days.append({"day": day, "times": selected_times})

        if not selected_days:
            QMessageBox.warning(self, "خطا", "لطفاً حداقل یک روز و ساعت را انتخاب کنید.")
            return

        if code not in self.teachers:
            self.teachers[code] = []

        for entry in selected_days:
            self.teachers[code].append({
                "name": name,
                "class": class_name,
                "course": course,
                "day": entry["day"],
                "times": entry["times"]
            })

        self.clear_inputs()
        self.update_table()
        QMessageBox.information(self, "موفق", f"استاد با کد {code} اضافه شد.")



    def edit_teacher(self):
        selected_row = self.teacher_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "خطا", "لطفاً یک سطر را برای ویرایش انتخاب کنید.")
            return

        teacher_id_item = self.teacher_table.item(selected_row, 0)
        if not teacher_id_item:
            QMessageBox.warning(self, "خطا", "اطلاعات استاد انتخاب شده معتبر نیست.")
            return

        teacher_id = teacher_id_item.text()
        teacher_entries = self.teachers.get(teacher_id, [])
        if not teacher_entries:
            QMessageBox.warning(self, "خطا", "اطلاعات استاد یافت نشد.")
            return

        current_name = self.teacher_table.item(selected_row, 1).text()
        current_class = self.teacher_table.item(selected_row, 2).text()
        current_course = self.teacher_table.item(selected_row, 3).text()
        current_day_time = self.teacher_table.item(selected_row, 4).text()

        name, ok1 = QInputDialog.getText(self, "ویرایش نام", "نام جدید استاد:", text=current_name)
        if not ok1:
            return

        class_name, ok2 = QInputDialog.getText(self, "ویرایش کلاس", "کلاس جدید استاد:", text=current_class)
        if not ok2:
            return

        course, ok3 = QInputDialog.getText(self, "ویرایش نام درس", "نام درس جدید:", text=current_course)
        if not ok3:
            return

        for entry in teacher_entries:
            day_time_str = f"{entry['day']}: {', '.join(entry['times'])}"
            if day_time_str == current_day_time:
                entry["name"] = name
                entry["class"] = class_name
                entry["course"] = course

        self.update_table()
        QMessageBox.information(self, "ویرایش موفق", f"اطلاعات استاد با کد {teacher_id} ویرایش شد.")

    def delete_teacher(self):
        selected_row = self.teacher_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "خطا", "لطفاً یک سطر را برای حذف انتخاب کنید.")
            return

        teacher_id_item = self.teacher_table.item(selected_row, 0)
        if not teacher_id_item:
            QMessageBox.warning(self, "خطا", "اطلاعات استاد انتخاب شده معتبر نیست.")
            return

        teacher_id = teacher_id_item.text()
        day_time_item = self.teacher_table.item(selected_row, 4)
        if not day_time_item:
            QMessageBox.warning(self, "خطا", "اطلاعات روز و ساعت انتخاب شده معتبر نیست.")
            return

        selected_day_time = day_time_item.text()
        if teacher_id in self.teachers:
            updated_entries = [
                entry for entry in self.teachers[teacher_id]
                if f"{entry['day']}: {', '.join(entry['times'])}" != selected_day_time
            ]
            if updated_entries:
                self.teachers[teacher_id] = updated_entries
            else:
                del self.teachers[teacher_id]

        self.update_table()
        QMessageBox.information(self, "حذف موفق", f"اطلاعات مربوط به استاد با کد {teacher_id} حذف شد.")

    def update_table(self):
        self.teacher_table.setRowCount(0)
        for code, entries in self.teachers.items():
            for entry in entries:
                row_position = self.teacher_table.rowCount()
                self.teacher_table.insertRow(row_position)
                self.teacher_table.setItem(row_position, 0, QTableWidgetItem(code))
                self.teacher_table.setItem(row_position, 1, QTableWidgetItem(entry["name"]))
                self.teacher_table.setItem(row_position, 2, QTableWidgetItem(entry["class"]))
                self.teacher_table.setItem(row_position, 3, QTableWidgetItem(entry["course"]))
                self.teacher_table.setItem(row_position, 4, QTableWidgetItem(f"{entry['day']}: {', '.join(entry['times'])}"))

    def clear_inputs(self):
        self.name_input.clear()
        self.code_input.clear()
        self.class_input.clear()
        self.subject_input.clear()
        for day, (day_checkbox, time_checkboxes) in self.days_time_checkboxes.items():
            day_checkbox.setChecked(False)
            for time in time_checkboxes:
                time.setChecked(False)

    def start_receiving_data(self):
        if self.ser:
            self.timer = QTimer(self)
            self.timer.timeout.connect(self.receive_data)
            self.timer.start(1000)  # 1 ثانیه


    def receive_data(self):
        if self.ser.in_waiting > 0:
            line = self.ser.readline().decode('utf-8').strip()
            self.process_data(line)



    def process_data(self, line):
        try:
            print(f"داده دریافتی: {line}")  # برای دیباگ

            # نادیده گرفتن پیام‌های غیرضروری یا خالی
            if line.startswith("LoRa Receiver initialized.") or not line.strip():
                return

            # دریافت کد از خط ورودی
            parts = line.split('*')
            if len(parts) == 2:
                class_code = parts[0].strip()  # کد کلاس (E101)
                teacher_code = parts[1].strip()  # کد استاد (101)
                print(f"کد کلاس: {class_code}, کد استاد: {teacher_code}")  # برای دیباگ

                # دریافت زمان و تاریخ فعلی
                now = datetime.now()
                time_str = now.strftime("%Y-%m-%d %H:%M:%S")
                current_day = self.days_of_week[now.weekday()]  # روز فعلی
                current_time = now.strftime("%H:%M")  # زمان فعلی

                # بررسی وجود استاد با کد استاد
                if teacher_code in self.teachers:
                    teacher_data = self.teachers[teacher_code]
                    status = "🔵"  # وضعیت پیش‌فرض

                    print(f"روز فعلی: {current_day}, زمان فعلی: {current_time}")
                    print(f"داده‌های استاد: {teacher_data}")

                    # بررسی تطابق روز و زمان
                    match_found = False
                    for record in teacher_data:
                        if record['day'] == current_day and record['class'] == class_code:
                            if any(self.is_time_in_slot(current_time, slot) for slot in record['times']):
                                match_found = True
                                # چک کردن زمان آخرین حضور
                                last_time = self.last_attendance.get(teacher_code)
                                if last_time:
                                    elapsed_time = now - last_time
                                    if timedelta(hours=1) <= elapsed_time <= timedelta(hours=1, minutes=30):
                                        status = "✔️"  # تیک آبی برای حضور در بازه یک ساعت تا یک ساعت و نیم
                                    elif elapsed_time < timedelta(hours=1):  # اگر کمتر از یک ساعت باشد
                                        print(f"استاد {teacher_code} زودتر از یک ساعت حضور مجدد ثبت کرده است.")
                                        return
                                # به‌روزرسانی زمان آخرین حضور
                                self.last_attendance[teacher_code] = now
                                break

                    if not match_found:
                        status = "❌"  # علامت عدم حضور

                    # افزودن رکورد به جدول حضور
                    row_position = self.attendance_table.rowCount()
                    self.attendance_table.insertRow(row_position)
                    self.attendance_table.setItem(row_position, 0, QTableWidgetItem(time_str))
                    self.attendance_table.setItem(row_position, 1, QTableWidgetItem(teacher_code))
                    self.attendance_table.setItem(row_position, 2, QTableWidgetItem(record['name']))
                    self.attendance_table.setItem(row_position, 3, QTableWidgetItem(class_code))
                    self.attendance_table.setItem(row_position, 4, QTableWidgetItem(current_day))
                    self.attendance_table.setItem(row_position, 5, QTableWidgetItem(status))

                    # ثبت رکورد در فایل
                    self.save_attendance_to_file()
                else:
                    print(f"کد موجود نیست: {teacher_code}")
                    print(f"کدهای ذخیره‌شده: {list(self.teachers.keys())}")
                    QMessageBox.warning(self, "خطا", f"استادی با این کد ({teacher_code}) وجود ندارد.")
            else:
                print("فرمت داده دریافتی نامعتبر است.")
                QMessageBox.warning(self, "خطا", "فرمت داده دریافتی نامعتبر است. فرمت صحیح: <کد کلاس>*<کد استاد>")
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطایی رخ داده است: {e}")


    def load_teachers_from_file(self):
        try:
            with open('teachers.json', 'r', encoding='utf-8') as file:
                self.teachers = json.load(file)
                self.fix_days_structure()
                self.update_table()
        except FileNotFoundError:
            self.teachers = {}  # فایل وجود ندارد، دیکشنری خالی ایجاد می‌شود
        except json.JSONDecodeError:
            QMessageBox.warning(self, "خطا", "خطای تجزیه فایل JSON: داده نامعتبر است.")

    def fix_days_structure(self):
        """اصلاح ساختار days اگر به اشتباه به عنوان لیست بارگذاری شده باشد."""
        for teacher in self.teachers.values():
            if 'days' in teacher:  # بررسی وجود کلید 'days'
                if isinstance(teacher['days'], list):
                    teacher['days'] = {day: teacher['days'] for day in self.days_of_week}

    def save_teachers_to_file(self):
        """ذخیره‌سازی اطلاعات اساتید در فایل JSON."""
        try:
            with open('teachers.json', 'w', encoding='utf-8') as file:
                json.dump(self.teachers, file, ensure_ascii=False, indent=4)
        except Exception as e:
            QMessageBox.warning(self, "خطا", f"خطا در ذخیره‌سازی فایل: {str(e)}")
    def save_attendance_to_file(self):
        data = []
        for row in range(self.attendance_table.rowCount()):
            row_data = []
            for col in range(self.attendance_table.columnCount()):
                item = self.attendance_table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)
        with open('attendance.csv', 'w', encoding='utf-8') as file:
            for row in data:
                file.write(','.join(row) + '\n')

    def save_attendance_to_excel(self):
        """ذخیره جدول حضور در فایل Excel و اضافه کردن به فایل موجود"""
        try:
            file_name = "attendance_records.xlsx"
            # بررسی اینکه فایل موجود است یا نه
            try:
                workbook = load_workbook(file_name)
                sheet = workbook.active
            except FileNotFoundError:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Attendance"
                # اضافه کردن عناوین ستون‌ها
                sheet.append(["تاریخ (شمسی)", "کد", "نام", "کلاس", "روز", "زمان کلاس"])

            # اضافه کردن داده‌های جدید به فایل
            for row in range(self.attendance_table.rowCount()):
                # دریافت تاریخ شمسی
                jalali_date = jdatetime.datetime.now().strftime("%Y/%m/%d")
                code = self.attendance_table.item(row, 1).text() if self.attendance_table.item(row, 1) else ""
                name = self.attendance_table.item(row, 2).text() if self.attendance_table.item(row, 2) else ""
                class_name = self.attendance_table.item(row, 3).text() if self.attendance_table.item(row, 3) else ""
                day = self.attendance_table.item(row, 4).text() if self.attendance_table.item(row, 4) else ""
                time_slot = self.attendance_table.item(row, 5).text() if self.attendance_table.item(row, 5) else ""
                sheet.append([jalali_date, code, name, class_name, day, time_slot])

            # ذخیره فایل
            workbook.save(file_name)
            QMessageBox.information(self, "موفقیت", f"جدول حضور با موفقیت به فایل {file_name} اضافه شد.")
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در ذخیره فایل Excel:\n{e}")

    def is_time_in_slot(self, current_time, time_slot):
        if '-' not in time_slot:
            print(f"فرمت زمان نامعتبر: {time_slot}")
            return False

        start_time_str, end_time_str = time_slot.split('-')
        start_time = datetime.strptime(start_time_str.strip(), "%H:%M").time()
        end_time = datetime.strptime(end_time_str.strip(), "%H:%M").time()
        current_time_obj = datetime.strptime(current_time, "%H:%M").time()

        print(f"بررسی زمان: {current_time_obj} در بازه {start_time} تا {end_time}")

        return start_time <= current_time_obj <= end_time
    def validate_days_and_times(self):
        # بررسی انتخاب حداقل یک روز و زمان
        for day, (day_checkbox, time_checkboxes) in self.days_time_checkboxes.items():
            if day_checkbox.isChecked():  # اگر روز انتخاب شده باشد
                for time_checkbox in time_checkboxes:
                    if time_checkbox.isChecked():  # اگر یکی از زمان‌های این روز انتخاب شده باشد
                        return True
        return False

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = TeacherManager()
    window.show()
    sys.exit(app.exec_())
